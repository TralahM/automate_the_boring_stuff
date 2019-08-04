#! python
'''Read and Manipulate Excel Spreadsheets, Create Charts, Formulas,etc
Author: Tralah M Brian
Company: TralahTek LLC.
'''
import logging
from argparse import ArgumentParser
import openpyxl

# additionally pass filename keyword arg to log to file instead of screen
logging.basicConfig(level=logging.DEBUG,
                    format="%(asctime)s - %(levelname)s - %(message)s")


def open_wb(fl):
    wb = openpyxl.load_workbook(fl)
    print("Sheets in this workbook:\n %s" % " ".join(wb.get_sheet_names()))
    sheets = [wb.get_sheet_by_name(name) for name in wb.get_sheet_names()]
    print(sheets[0]['A1'].value)

    # cell=sheet["A2"], cell.row,cell.column,cell.value,cell.coordinate
    # sheet.cell(row=1,column=2)
    # Shape of sheets --> sheet.get_highest_row(),sheet.get_highest_column()
    # openpyxl.cell.column_index_from_string()
    # openpyxl.cell.get_column_letter()
    # sheet['A1':'C4']


if __name__ == '__main__':
    ps = ArgumentParser()
    ps.add_argument('-f', '--file', action='store', dest='fl',
                    help=" The Excel Spreadsheet to open.")
    args = ps.parse_args()
    logging.info("Opening Excel Spreadsheet: " + args.fl)
    open_wb(args.fl)
